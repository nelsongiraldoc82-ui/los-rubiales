#!/usr/bin/env python3
"""
Script para generar Excel de registros de huéspedes del Hotel Los Rubiales
"""
import sys
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colores del hotel
HOTEL_GREEN = "1a5f2a"
HEADER_FILL = "1F4E79"
ALT_ROW_FILL = "F5F5F5"
WHITE = "FFFFFF"

def format_date(date_str):
    """Formatear fecha ISO a formato legible"""
    if not date_str:
        return '-'
    try:
        dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        return dt.strftime('%d/%m/%Y %H:%M')
    except:
        return date_str

def format_date_short(date_str):
    """Formatear fecha corta"""
    if not date_str:
        return '-'
    try:
        dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        return dt.strftime('%d/%m/%Y')
    except:
        return date_str

def generate_excel(data, output_path, language='es'):
    """Generar el Excel con los datos proporcionados"""

    # Textos según idioma
    texts = {
        'es': {
            'title': 'Registro de Huéspedes - Hotel Rural Los Rubiales',
            'apartment': 'Apartamento',
            'checkin': 'Entrada',
            'checkout': 'Salida',
            'status': 'Estado',
            'active': 'Activo',
            'checked_out': 'Finalizado',
            'guest_name': 'Nombre',
            'guest_lastname': 'Apellidos',
            'document_type': 'Tipo Doc.',
            'document_number': 'Número Doc.',
            'nationality': 'Nacionalidad',
            'email': 'Email',
            'phone': 'Teléfono',
            'main_guest': 'Principal',
            'notes': 'Notas',
            'yes': 'Sí',
            'no': 'No',
            'total': 'Total',
            'registrations': 'registros',
            'guests': 'huéspedes',
            'generated': 'Generado el',
        },
        'en': {
            'title': 'Guest Registration - Hotel Rural Los Rubiales',
            'apartment': 'Apartment',
            'checkin': 'Check-in',
            'checkout': 'Check-out',
            'status': 'Status',
            'active': 'Active',
            'checked_out': 'Checked Out',
            'guest_name': 'First Name',
            'guest_lastname': 'Last Name',
            'document_type': 'Doc. Type',
            'document_number': 'Doc. Number',
            'nationality': 'Nationality',
            'email': 'Email',
            'phone': 'Phone',
            'main_guest': 'Main Guest',
            'notes': 'Notes',
            'yes': 'Yes',
            'no': 'No',
            'total': 'Total',
            'registrations': 'registrations',
            'guests': 'guests',
            'generated': 'Generated on',
        }
    }

    t = texts.get(language, texts['es'])

    # Crear workbook
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Registros' if language == 'es' else 'Registrations'

    # Estilos
    title_font = Font(name='Times New Roman', size=18, bold=True, color=HOTEL_GREEN)
    header_font = Font(name='Times New Roman', size=11, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
    cell_font = Font(name='Times New Roman', size=10)
    alt_fill = PatternFill(start_color=ALT_ROW_FILL, end_color=ALT_ROW_FILL, fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Título
    sheet['B2'] = t['title']
    sheet['B2'].font = title_font
    sheet.merge_cells('B2:L2')

    # Fecha de generación
    sheet['B3'] = f"{t['generated']}: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sheet['B3'].font = Font(name='Times New Roman', size=10, italic=True, color='666666')

    # Resumen
    total_regs = len(data)
    total_guests = sum(len(r.get('guests', [])) for r in data)
    active_count = sum(1 for r in data if r.get('status') == 'active')

    sheet['B5'] = f"{t['total']}: {total_regs} {t['registrations']}, {total_guests} {t['guests']} ({active_count} {t['active']})"
    sheet['B5'].font = Font(name='Times New Roman', size=10, bold=True)
    sheet.merge_cells('B5:L5')

    # Encabezados (fila 7)
    headers = [
        '#', t['apartment'], t['checkin'], t['checkout'], t['status'],
        t['guest_name'], t['guest_lastname'], t['document_type'],
        t['document_number'], t['nationality'], t['email'], t['phone'],
        t['main_guest'], t['notes']
    ]

    for col, header in enumerate(headers, 2):
        cell = sheet.cell(row=7, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Datos
    row_num = 8
    for reg_idx, reg in enumerate(data, 1):
        guests = reg.get('guests', [])
        if not guests:
            guests = [{}]  # Fila vacía si no hay huéspedes

        apartment_name = reg.get('apartment', {}).get('name', 'N/A')
        checkin = format_date(reg.get('checkInDate', ''))
        checkout = format_date_short(reg.get('checkOutDate')) if reg.get('checkOutDate') else '-'
        status_text = t['active'] if reg.get('status') == 'active' else t['checked_out']
        notes = reg.get('notes', '')

        for guest_idx, guest in enumerate(guests):
            # Alternar color de fila
            row_fill = alt_fill if (row_num - 8) % 2 == 0 else None

            # Número de registro (solo en primera fila del grupo)
            sheet.cell(row=row_num, column=2).value = reg_idx if guest_idx == 0 else ''
            sheet.cell(row=row_num, column=3).value = apartment_name if guest_idx == 0 else ''
            sheet.cell(row=row_num, column=4).value = checkin if guest_idx == 0 else ''
            sheet.cell(row=row_num, column=5).value = checkout if guest_idx == 0 else ''
            sheet.cell(row=row_num, column=6).value = status_text if guest_idx == 0 else ''

            # Datos del huésped
            sheet.cell(row=row_num, column=7).value = guest.get('firstName', '')
            sheet.cell(row=row_num, column=8).value = guest.get('lastName', '')
            sheet.cell(row=row_num, column=9).value = guest.get('documentType', '')
            sheet.cell(row=row_num, column=10).value = guest.get('documentNumber', '')
            sheet.cell(row=row_num, column=11).value = guest.get('nationality', '')
            sheet.cell(row=row_num, column=12).value = guest.get('email', '')
            sheet.cell(row=row_num, column=13).value = guest.get('phone', '')
            sheet.cell(row=row_num, column=14).value = t['yes'] if guest.get('isMainGuest') else t['no']
            sheet.cell(row=row_num, column=15).value = notes if guest_idx == 0 else ''

            # Aplicar estilos
            for col in range(2, 16):
                cell = sheet.cell(row=row_num, column=col)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = center_align if col < 7 else left_align
                if row_fill:
                    cell.fill = row_fill

            row_num += 1

    # Ajustar anchos de columna
    column_widths = {
        'B': 5,   # #
        'C': 15,  # Apartamento
        'D': 18,  # Entrada
        'E': 12,  # Salida
        'F': 12,  # Estado
        'G': 12,  # Nombre
        'H': 15,  # Apellidos
        'I': 10,  # Tipo Doc
        'J': 15,  # Número Doc
        'K': 12,  # Nacionalidad
        'L': 25,  # Email
        'M': 15,  # Teléfono
        'N': 10,  # Principal
        'O': 30,  # Notas
    }

    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    # Congelar panel
    sheet.freeze_panes = 'B8'

    # Guardar
    wb.save(output_path)
    return output_path

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python generate_registrations_xlsx.py <output_path> <language> [json_data]")
        sys.exit(1)

    output_path = sys.argv[1]
    language = sys.argv[2]

    # Leer datos de stdin o argumento
    if len(sys.argv) > 3:
        json_data = sys.argv[3]
    else:
        json_data = sys.stdin.read()

    try:
        data = json.loads(json_data)
        generate_excel(data, output_path, language)
        print(f"Excel generated successfully: {output_path}")
    except Exception as e:
        print(f"Error generating Excel: {e}", file=sys.stderr)
        sys.exit(1)
